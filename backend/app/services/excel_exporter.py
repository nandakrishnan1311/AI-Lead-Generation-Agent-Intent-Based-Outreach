"""
services/excel_exporter.py — Exports all leads from SQLite to a formatted Excel file.

Sheet layout:
  - Sheet 1: Full lead list with all fields
  - Sheet 2: Summary statistics (top companies by score, source breakdown)
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.config import REPORT_OUTPUT_PATH
from app.database.models import Lead
from app.utils.logger import get_logger

logger = get_logger(__name__)

# Colour palette
HEADER_FILL  = PatternFill("solid", fgColor="1A1A2E")   # dark navy
ALT_FILL     = PatternFill("solid", fgColor="F0F4FF")   # light blue-tint
HIGH_FILL    = PatternFill("solid", fgColor="D4EDDA")   # green (score ≥ 7)
MED_FILL     = PatternFill("solid", fgColor="FFF3CD")   # amber (score 4–6)
LOW_FILL     = PatternFill("solid", fgColor="F8D7DA")   # red (score < 4)
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT   = Font(bold=True, size=13, color="1A1A2E")

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COLUMNS = [
    ("Company Name",    "company",    30),
    ("Contact Name",    "contact",    22),
    ("Title",           "title",      18),
    ("LinkedIn URL",    "linkedin",   40),
    ("Company Website", "website",    30),
    ("Signal Summary",  "signal",     55),
    ("Signal Source URL","signal_url",40),
    ("Intent Score",    "score",      13),
    ("Reasoning",       "reasoning",  50),
    ("Date Found",      "date_found", 18),
]


def _score_fill(score: float) -> PatternFill:
    if score >= 7:
        return HIGH_FILL
    if score >= 4:
        return MED_FILL
    return LOW_FILL


def export_leads(db: Session, output_path: Path | None = None) -> Path:
    """Export all leads to Excel; return the output path."""
    path = output_path or REPORT_OUTPUT_PATH
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    leads: list[Lead] = db.query(Lead).order_by(Lead.score.desc()).all()
    logger.info("Exporting %d leads to %s", len(leads), path)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Lead List ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Lead List"
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"AI Lead Generation Report  •  Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 22

    # Freeze header rows
    ws.freeze_panes = "A3"

    # Data rows
    for row_idx, lead in enumerate(leads, start=3):
        fill = _score_fill(lead.score or 0)
        alt = ALT_FILL if row_idx % 2 == 0 else None

        date_str = lead.date_found.strftime("%Y-%m-%d %H:%M") if lead.date_found else ""

        values = [
            lead.company, lead.contact, lead.title, lead.linkedin,
            lead.website, lead.signal, lead.signal_url,
            round(lead.score or 0, 1), lead.reasoning, date_str,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            # Score column gets coloured fill; others get alternating
            if col_idx == 8:
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif alt:
                cell.fill = alt

        ws.row_dimensions[row_idx].height = 40

    # Auto-filter
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}2"

    # ── Sheet 2: Summary ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15

    headers2 = [("Metric", "Value")]
    data2 = [
        ("Total Leads",            len(leads)),
        ("High Intent (score ≥ 7)", sum(1 for l in leads if (l.score or 0) >= 7)),
        ("Medium Intent (4–6)",    sum(1 for l in leads if 4 <= (l.score or 0) < 7)),
        ("Low Intent (< 4)",       sum(1 for l in leads if (l.score or 0) < 4)),
        ("Avg Intent Score",       round(sum(l.score or 0 for l in leads) / max(len(leads), 1), 2)),
        ("Contacts Found",         sum(1 for l in leads if l.contact)),
        ("LinkedIn URLs Found",    sum(1 for l in leads if l.linkedin)),
        ("Report Generated (UTC)", datetime.utcnow().strftime("%Y-%m-%d %H:%M")),
    ]

    row = 1
    for (h1, h2) in headers2:
        ws2.cell(row=row, column=1, value=h1).font = HEADER_FONT
        ws2.cell(row=row, column=1).fill = HEADER_FILL
        ws2.cell(row=row, column=2, value=h2).font = HEADER_FONT
        ws2.cell(row=row, column=2).fill = HEADER_FILL
        row += 1

    for label, value in data2:
        c1 = ws2.cell(row=row, column=1, value=label)
        c2 = ws2.cell(row=row, column=2, value=value)
        c1.border = BORDER
        c2.border = BORDER
        if row % 2 == 0:
            c1.fill = ALT_FILL
            c2.fill = ALT_FILL
        row += 1

    wb.save(path)
    logger.info("Excel report saved: %s", path)
    return path
